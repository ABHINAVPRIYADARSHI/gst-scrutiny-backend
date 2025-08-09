import os
import pandas as pd
import datetime
from glob import glob
from collections import defaultdict

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.globals.constants import parse_month_year, late_fee_headers, parse_month

financial_year_2021_22 = "2021-22"
special_months = [11, 12, 1, 2, 3]


async def generate_gstr1_merged(input_dir, output_dir):
    print(f"[GSTR-1_merged] Started execution of method generate_gstr1_merged for: {input_dir}")
    final_result_points = {}
    output_path = None

    try:
        excel_files = sorted(glob(os.path.join(input_dir, "*.xlsx")))
        if not excel_files:
            raise FileNotFoundError("No GSTR-1 Excel files found in input directory.")
        print(f"[GSTR-1] Found {len(excel_files)} Excel files.")

        header_map = {}  # sheet_name -> [header_row_1, header_row_2]
        sheet_data = defaultdict(list)
        list_of_dfs = []

        for file_index, file_path in enumerate(excel_files):
            wb = load_workbook(file_path, data_only=True)
            print(f"Processing file: {os.path.basename(file_path)}")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                if sheet_name.lower().strip() == "read me":
                    rows = [
                        row for row in ws.iter_rows(min_row=4, max_row=10, min_col=2, max_col=3, values_only=True)
                        if any(cell is not None for cell in row)
                    ]
                    if rows:
                        df = pd.DataFrame(rows, columns=["Field", "Value"])
                        list_of_dfs.append(df)
                else:
                    # Capture headers only from the first file
                    if file_index == 0:
                        header_rows = []
                        for row_idx in [2, 3]:  # 3rd and 4th rows (0-indexed)
                            row_values = [cell.value if cell.value is not None else "" for cell in ws[row_idx + 1]]
                            header_rows.append(row_values)
                        header_map[sheet_name] = header_rows

                    # Extract data from row 5 onward
                    data_rows = []
                    for row in ws.iter_rows(min_row=5, values_only=True):
                        if all(cell is None for cell in row):
                            continue
                        data_rows.append(row)

                    if data_rows:
                        df = pd.DataFrame(data_rows)
                        sheet_data[sheet_name].append(df)
                    else:
                        sheet_data[sheet_name]  # Ensure key exists even if empty

        # ✅ Late Fee Calculation
        late_fee_records = calculate_late_fee(list_of_dfs)
        total_late_fee = sum(row[-1] for row in late_fee_records if len(row) >= 7)
        print(f"[GSTR-1_merged] Total late fee: {total_late_fee}")
        final_result_points["result_point_12_total_late_fee_gstr1"] = total_late_fee

        # ✅ Write Final Output Excel
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "GSTR-1_merged.xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, header_rows in header_map.items():
                df_list = sheet_data.get(sheet_name)
                if df_list:
                    combined_df = pd.concat(df_list, ignore_index=True)
                    combined_df.to_excel(
                        writer,
                        sheet_name=sheet_name + "_merged",
                        header=False,
                        index=False,
                        startrow=len(header_rows),
                        startcol=0
                    )
                    for header in header_rows:
                        writer.sheets[sheet_name + "_merged"].append(header)

            # ✅ Insert Late Fee Records
            late_fee_df = pd.DataFrame(late_fee_records, columns=late_fee_headers)
            late_fee_df.to_excel(
                writer,
                sheet_name="Late Fee Record",
                header=True,
                index=False
            )
            writer.sheets["Late Fee Record"].insert_rows(0, 1)  # Add a blank row at top

        # ✅ Format Sheets
        wb = load_workbook(output_path)
        format_workbook_sheets(wb, col_width=25)

        # ✅ Save File
        wb.save(output_path)
        print(f"✅ GSTR-1 merged Excel saved to: {output_path}")
        return output_path, final_result_points
    except Exception as e:
        print(f"[GSTR-1_merged]: ❌ Error while merging GSTR-1 files. {e}")
        return output_path, final_result_points


#  Parameter 12 of ASMT-10 report
def calculate_late_fee(list_of_dataframes):
    print(f"[GSTR-1_merged] Starting late fee calculation, number of months: {len(list_of_dataframes)}")
    records = []
    for entry in list_of_dataframes:
        try:
            financial_year = entry.iloc[0, 1]
            return_month = entry.iloc[1, 1]
            print(f"Calculating late fee for: {(return_month, financial_year)}")

            filing_date_str = entry.iloc[6, 1]
            filing_date = datetime.datetime.strptime(filing_date_str, "%d-%m-%Y").date()

            # Due date for FY 2021-22 is 13th of next month for return months Nov, Dec, Jan, Feb, Mar
            # else Due date = 11th of next month
            if financial_year == financial_year_2021_22 and parse_month(return_month) in special_months:
                print(f"[GSTR-1] Late fee calculation: Setting due day as 13th of next month for month: {return_month}, year: {financial_year}")
                day = 13
            else:
                print(f"[GSTR-1] Late fee calculation: Setting due day as 11th of next month for month: {return_month}, year: {financial_year}")
                day = 11
            return_month_date = parse_month_year(return_month, financial_year)
            due_date = (return_month_date + relativedelta(months=1)).replace(day=day)
            print(f"Due date: {due_date}")
            days_late = max((filing_date - due_date).days, 0)
            calculated_late_fee = 100 * days_late
            late_fee_applicable = min(calculated_late_fee, 5000)

            records.append([
                financial_year,
                return_month,
                filing_date.strftime("%d-%m-%Y"),
                due_date.strftime("%d-%m-%Y"),
                days_late,
                calculated_late_fee,
                late_fee_applicable
            ])
        except Exception as e:
            print(f"[GSTR-1_merged]: ❌ Error calculating late fee for: {(return_month, financial_year)}")
            print(e)
    return records


def format_workbook_sheets(wb, col_width=25):
    for ws in wb.worksheets:
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_width

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True)
                    if row_idx == 1:
                        cell.font = Font(bold=True)