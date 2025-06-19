from openpyxl import Workbook
from utils.globals.constants import estimatedITCReversal, result_point_9, result_point_14, result_point_3, \
    result_point_7, result_point_19, result_point_20
from utils.globals.constants import diffInRCMPayment
from utils.globals.constants import diffInRCM_ITC
from .gstr3b_merged_reader import gstr3b_merged_reader
from openpyxl.styles import Alignment


async def generate_gstr3b_merged_analysis(gstin):
    print(" === Starting execution of file gstr3B_analysis.py ===")
    final_result_points = {}
    output_path = f"reports/{gstin}/GSTR-3B_analysis.xlsx"
    try:
        valuesFrom3b = await gstr3b_merged_reader(gstin)
        # Create workbook and sheet
        wb = Workbook()
        # Define sheet names for each pair
        sheet_names = ["Estimated Prop. ITC reversal", "Diif. in RCM ITC", "Diff. in RCM Payment"]
        # Define your data pairs
        data_pairs = [
            ("Estimated Proportionate ITC Reversal", valuesFrom3b.get(estimatedITCReversal)),
            ("Difference in RCM ITC", valuesFrom3b.get(diffInRCM_ITC)),
            ("Difference in RCM payment", valuesFrom3b.get(diffInRCMPayment))
        ]
        # Fill each sheet with one row
        for idx, (sheet_name, (label, value)) in enumerate(zip(sheet_names, data_pairs)):
            if idx == 0:
                # Use default sheet for first item and rename it
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Optional: set column width
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20

            # Write label and rounded value
            ws.cell(row=1, column=1, value=label).alignment = Alignment(wrap_text=True)
            ws.cell(row=1, column=2, value=round(value, 2))

        # Save the workbook
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        print(" === ✅ Returning after successful execution of file gstr3b_analysis.py ===")
        final_result_points[result_point_3] = valuesFrom3b.get(diffInRCM_ITC)
        final_result_points[result_point_7] = valuesFrom3b.get(estimatedITCReversal)
        final_result_points[result_point_14] = valuesFrom3b.get("table_3_1_a+c+e_taxable_value")
        final_result_points[result_point_19] = valuesFrom3b.get("sum_table_4A_row_4")
        final_result_points[result_point_20] = valuesFrom3b.get("sum_table_4A_row_1")
        return final_result_points
    except Exception as e:
        print(f"[GSTR-3B Analysis] ❌ Error during analysis: {e}")
        return final_result_points
