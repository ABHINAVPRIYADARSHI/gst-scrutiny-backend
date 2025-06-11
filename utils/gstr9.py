import os
import pandas as pd
from glob import glob
from openpyxl import Workbook
import pdfplumber
from tabulate import tabulate
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Define how many rows to skip per table before setting headers. Don't modify the structure else
#table_position_in_cleaned_tables will have to be changed accordingly.
table_header_rows_skip = {
    0: -1,
    1 : 4, # Table 4
    2 : 0, # Table 4
    3 : 4, # Table 5
    4 : -1, # Table 5
    # 5 : 5, not required
    # 6: 0,  not required
    # 7 : 3, # Table 6
    8 : 3,  # Table 7
    # 9 : 0,  Part of Table 7 not required
    10 : 3, # Table 8
    11 : 3, # Table 9
}

table_position_in_cleaned_tables = {
    "Table_4_part_I" : 1,
    "Table_4_part_II" : 2,
    "Table_5_part_I" : 3,
    "Table_5_part_II" : 4,
    "Table_7_part_I" : 5,
    "Table_8" : 6,
    "Table_9" : 7
}
cleaned_tables = []
all_tables = []

async def generate_gstr9(input_dir, output_dir):
    print("Extracting GSTR-9 from PDF in excel form...")
    pdf_files = glob(os.path.join(input_dir, "*.pdf"))
    if not pdf_files:
        raise FileNotFoundError("No PDF files found in input directory.")
        return
    print(f"Found {len(pdf_files)} PDF files.")

    # Read the only annual GSTR-9 return and extract tables
    with pdfplumber.open(pdf_files[0]) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                all_tables.append(table)

    # Clean and process specific tables
    for idx, skip_rows in table_header_rows_skip.items():
        if idx < len(all_tables):
            df = pd.DataFrame(all_tables[idx])
            df.columns = df.iloc[skip_rows]  # Set new header
            df = df.drop(index=range(skip_rows + 1))  # Drop the header rows
            df = df.reset_index(drop=True)  # Reset index
            cleaned_tables.append(df)
            print(f"Table no: {idx}")
            print(tabulate(df, tablefmt='grid', maxcolwidths=20))

    # Create a new workbook and sheet (always fresh)
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-9 vs GSTR-3B"
    row_cursor = 1  # Track row position in Excel
    output_file = os.path.join(output_dir, "GSTR-9 vs GSTR-3B.xlsx")

    # 1. Table_4_row_G
    table_4 = cleaned_tables[table_position_in_cleaned_tables["Table_4_part_I"]]
    ws.cell(row = row_cursor, column =1, value= table_4.iloc[6, 1])
    ws.cell(row = row_cursor, column =2, value= table_4.iloc[6, 2])
    row_cursor += 2

    # 2. Table_4_row_N
    table_4 = cleaned_tables[table_position_in_cleaned_tables["Table_4_part_II"]]
    ws.cell(row=row_cursor, column=1, value=table_4.iloc[7, 1])
    ws.cell(row=row_cursor, column=2, value=table_4.iloc[7, 2])
    row_cursor += 2

    # 3. Table_5_row_D
    table_5 = cleaned_tables[table_position_in_cleaned_tables["Table_5_part_I"]]
    ws.cell(row=row_cursor, column=1, value=table_5.iloc[4, 1])
    ws.cell(row=row_cursor, column=2, value=table_5.iloc[4, 2])
    row_cursor += 1

    # 4. Table_5_row_N
    table_5 = cleaned_tables[table_position_in_cleaned_tables["Table_5_part_II"]]
    ws.cell(row=row_cursor, column=1, value=table_5.iloc[9, 1])
    ws.cell(row=row_cursor, column=2, value=table_5.iloc[9, 2])
    row_cursor += 2

    # 5. Table_7_row_A
    table_7 = cleaned_tables[table_position_in_cleaned_tables["Table_7_part_I"]]
    # df = pd.DataFrame(table_7)
    sum = pd.to_numeric(table_7.iloc[0,2:], errors='coerce').sum(skipna=True)
    ws.cell(row=row_cursor, column=1, value="Interest Liability")
    ws.cell(row=row_cursor, column=2, value= sum)
    row_cursor += 1
    # 5. Table_7_row_C
    ws.cell(row=row_cursor, column=1, value="Proportionate Reversal")
    row_values = table_7.iloc[2, 2:].apply(lambda x: str(x).replace(',', '').strip())
    sum = pd.to_numeric(row_values, errors='coerce').sum(skipna=True)
    ws.cell(row=row_cursor, column=2, value= sum)
    row_cursor += 2

    # 6. Table_8_row_D
    table_8 = cleaned_tables[table_position_in_cleaned_tables["Table_8"]]
    ws.cell(row=row_cursor, column=1, value=table_8.iloc[3, 1])
    ws.cell(row=row_cursor, column=2, value=table_8.iloc[3, 2])
    ws.cell(row=row_cursor, column=3, value=table_8.iloc[3, 3])
    ws.cell(row=row_cursor, column=4, value=table_8.iloc[3, 4])
    ws.cell(row=row_cursor, column=5, value=table_8.iloc[3, 5])
    row_values = table_8.iloc[3, 2:].apply(lambda x: str(x).replace(',', '').strip())
    sum = pd.to_numeric(row_values, errors='coerce').sum(skipna=True)
    print(sum)
    if(sum < 0.00):
        ws.cell(row=row_cursor, column=6, value= sum)
        ws.cell(row=row_cursor, column=7, value="These values should be non-negative")
    row_cursor += 1
    # 7. Table_8_row_I
    ws.cell(row=row_cursor, column=1, value=table_8.iloc[8, 1])
    ws.cell(row=row_cursor, column=2, value=table_8.iloc[8, 2])
    ws.cell(row=row_cursor, column=3, value=table_8.iloc[8, 3])
    ws.cell(row=row_cursor, column=4, value=table_8.iloc[8, 4])
    ws.cell(row=row_cursor, column=5, value=table_8.iloc[8, 5])
    row_values = table_8.iloc[8, 2:].apply(lambda x: str(x).replace(',', '').strip())
    sum = pd.to_numeric(row_values, errors='coerce').sum(skipna=True)
    print(sum)
    if (sum < 0.00):
        ws.cell(row=row_cursor, column=6, value= sum)
        ws.cell(row=row_cursor, column=7, value="These values should be non-negative")
    row_cursor += 1

    # Apply wrap_text and fixed width to all cells in that column
    for col_idx in range(1, 8):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
        # Apply wrap_text to all cells in that column
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output_file)
    print(f"Excel file saved to: {output_file}")

def convert_to_number(value):
    try:
        # Clean value: remove commas, strip spaces
        cleaned = str(value).replace(',', '').strip()
        # Convert to float (handles int too)
        return float(cleaned)
    except (ValueError, TypeError):
        return value  # Leave as-is if not convertible