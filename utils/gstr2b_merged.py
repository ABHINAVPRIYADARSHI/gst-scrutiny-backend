import os
import pandas as pd
from glob import glob
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import copy

# Define header row ranges per sheet (0-indexed)
header_row_map = {
    "ITC Available": [4, 5, 6, 7, 8],
    "ITC not available": [4, 5, 6, 7, 8],
    "B2B": [3, 4, 5],
    "B2BA": [3, 4, 5, 6],
    "ECO": [3, 4, 5],
    "B2B-CDNR": [3, 4, 5],
    "B2B-CDNRA": [3, 4, 5, 6],
    "ISD": [3, 4, 5],
    "ISDA": [3, 4, 5, 6],
    "IMPG": [3, 4, 5],
    "IMPGSEZ": [3, 4, 5],
    # The following sheets are not available in GSTR-2B.
    # "ECOA": [3, 4, 5, 6],
    # "TDS": [3, 4, 5],
    # "TDSA": [3, 4, 5],
    # "TCS": [3, 4, 5]
}


def copy_header_with_styles(source_ws, target_ws, header_rows):
    """
    Copies specified header rows from source_ws to target_ws,
    preserving formatting, merged cells, column widths, and row heights.
    """
    header_rows = sorted(header_rows)

    # 1. Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    # 2. Copy styled header rows
    for target_row_idx, source_row_idx in enumerate(header_rows, start=1):
        source_row = source_ws[source_row_idx + 1]  # openpyxl is 1-indexed
        target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[source_row_idx + 1].height

        for col_idx, cell in enumerate(source_row, start=1):
            new_cell = target_ws.cell(row=target_row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                if cell.font: new_cell.font = copy.copy(cell.font)
                if cell.border: new_cell.border = copy.copy(cell.border)
                if cell.fill: new_cell.fill = copy.copy(cell.fill)
                if cell.number_format: new_cell.number_format = cell.number_format
                if cell.protection: new_cell.protection = copy.copy(cell.protection)
                if cell.alignment: new_cell.alignment = copy.copy(cell.alignment)

    # 3. Copy merged cells within header range
    header_row_set = set(r + 1 for r in header_rows)  # 1-based
    for merged_range in source_ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if all(row in header_row_set for row in range(min_row, max_row + 1)):
            try:
                mapped_min_row = header_rows.index(min_row - 1) + 1
                mapped_max_row = header_rows.index(max_row - 1) + 1
                target_ws.merge_cells(
                    start_row=mapped_min_row,
                    start_column=min_col,
                    end_row=mapped_max_row,
                    end_column=max_col
                )
            except ValueError:
                continue


async def generate_gstr2b_merged(input_dir, output_dir):
    print(f"[GSTR-2B] Started execution of method generate_gstr2b_merged for: {input_dir}")

    excel_files = sorted(glob(os.path.join(input_dir, "*.xlsx")))
    if not excel_files:
        raise FileNotFoundError("No Excel files found in the input directory.")

    print(f"Found {len(excel_files)} GSTR-2B Excel files.")

    sheet_data = defaultdict(list)
    readme_copy = None
    readme_done = False

    for file_idx, file_path in enumerate(excel_files):
        wb = load_workbook(file_path, data_only=True)
        print(f"Processing file: {os.path.basename(file_path)}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # We don't need to merge "read me" sheet
            # if sheet_name.lower().strip() == "read me":
            #     if not readme_done:
            #         readme_copy = ws
            #         readme_done = True
            #     continue

            if sheet_name not in header_row_map:
                print(f"Skipping unknown GSTR-2B sheet: {sheet_name}")
                continue

            header_rows = header_row_map[sheet_name]
            data_start_row = max(header_rows) + 1

            # Extract data rows
            data_rows = []
            for row in ws.iter_rows(min_row=data_start_row + 1, values_only=True):
                if all(cell is None for cell in row):
                    continue
                data_rows.append(row)

            if not data_rows:
                sheet_data[sheet_name]  # ensure key exists
                continue

            df = pd.DataFrame(data_rows)
            if not df.empty:
                sheet_data[sheet_name].append(df)

    # Prepare output
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "GSTR-2B_merged.xlsx")
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)

    for sheet_name in header_row_map:
        merged_ws = merged_wb.create_sheet(title=(sheet_name + "_merged")[:31])

        # Copy formatted header from the first file
        source_ws = load_workbook(excel_files[0], data_only=False)[sheet_name]
        header_rows = header_row_map[sheet_name]
        copy_header_with_styles(source_ws, merged_ws, header_rows)

        # Write stacked data
        df_list = sheet_data.get(sheet_name)
        if df_list:
            combined_df = pd.concat(df_list, ignore_index=True)
            for row in dataframe_to_rows(combined_df, index=False, header=False):
                merged_ws.append(row)

    merged_wb.save(output_path)
    print(f"✅ [GSTR-2B] merged Excel saved to: {output_path}")
    return output_path
