from openpyxl import Workbook
import pdfplumber
from tabulate import tabulate
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from .globals.constants import estimatedLiability, floating_zero
from .gstr9_reader import gstr9_reader
from .gstr3b_reader import gstr3b_reader


async def generate_gstr9_analysis(gstin):
    print(" === Starting execution of file gstr9_analysis.py ===")
    output_dir_of_GSTR_9_report = f"reports/{gstin}/GSTR-9/GSTR-9 vs GSTR-3B.xlsx"

    try:
        valuesFrom9 = await gstr9_reader(gstin)
        valuesFrom3b = await gstr3b_reader(gstin)
        # Create a new workbook and sheet (always fresh)
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-9 vs GSTR-3B"
        row_cursor = 1  # Track row position in Excel

        # 1. Table_4_row_G
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T4_G2")
        ws.cell(row=row_cursor, column=2, value=valuesFrom9["table4_G1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table4_G2"]))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_d")
        ws.cell(row=row_cursor, column=2, value=valuesFrom3b["table_3_1_D1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom3b["table_3_1_D2"]))
        row_cursor += 3

        # 2. Table_4_row_N
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T4_N2")
        ws.cell(row=row_cursor, column=2, value=valuesFrom9["table4_N1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table4_N2"]))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_(cell a + c)")
        ws.cell(row=row_cursor, column=3, value=valuesFrom3b["sum_table_3_1_A1_and_C1"])
        row_cursor += 3

        # 3. Table_5_row_D
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T5_D2")
        ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_5_D1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_5_D2"]))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_c")
        ws.cell(row=row_cursor, column=2, value=valuesFrom3b["table_3_1_C1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom3b["table_3_1_C2"]))
        row_cursor += 3

        # 4. Table_5_row_N
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T5_N2")
        ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_5_N1"])
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_5_N2"]))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_Sum_(a+b+c+d+e)")
        ws.cell(row=row_cursor, column=3, value=valuesFrom3b["sum_table_3_1_A1_to_E1"])
        row_cursor += 3

        # 5. Table_6_row_H
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T6_H (ITC reclaimed)")
        ws.cell(row=row_cursor, column=2, value="Interest Liability on amount: ")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["sum_table6_row_H"])
        row_cursor += 3

        # 6. Table_7_row_A
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T7_A (As per Rule 37)")
        ws.cell(row=row_cursor, column=2, value="Interest Liability on amount: ")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["sum_table7_row_A"])
        row_cursor += 3

        # 7. Table_7_row_C
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T7_C (As per Rule 42)")
        ws.cell(row=row_cursor, column=2, value=" ITC reversal as per GSTR-9")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["sum_table7_row_C"])
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged")
        ws.cell(row=row_cursor, column=2, value="Estimated Liability as per GSTR-3B")
        ws.cell(row=row_cursor, column=3, value=valuesFrom3b[estimatedLiability])
        row_cursor += 3

        # 8. Table_8_row_D
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T8_D")
        ws.cell(row=row_cursor, column=2, value=convert_to_number(valuesFrom9["table_8_D1"]))
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_8_D2"]))
        ws.cell(row=row_cursor, column=4, value=convert_to_number(valuesFrom9["table_8_D3"]))
        ws.cell(row=row_cursor, column=5, value=convert_to_number(valuesFrom9["table_8_D4"]))
        ws.cell(row=row_cursor, column=6, value=convert_to_number(valuesFrom9["table_8_D5"]))
        ws.cell(row=row_cursor, column=7, value=valuesFrom9["sum_table8_row_D"])
        # if valuesFrom9["sum_table8_row_D"] < floating_zero:
        ws.cell(row=row_cursor, column=8, value="None of these values should be negative")
        row_cursor += 3

        # 9. Table_8_row_I
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T8_I")
        ws.cell(row=row_cursor, column=2, value=convert_to_number(valuesFrom9["table_8_I1"]))
        ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_8_I2"]))
        ws.cell(row=row_cursor, column=4, value=convert_to_number(valuesFrom9["table_8_I3"]))
        ws.cell(row=row_cursor, column=5, value=convert_to_number(valuesFrom9["table_8_I4"]))
        ws.cell(row=row_cursor, column=6, value=convert_to_number(valuesFrom9["table_8_I5"]))
        ws.cell(row=row_cursor, column=7, value=valuesFrom9["sum_table8_row_I"])
        # if valuesFrom9["sum_table8_row_I"] < 0.00:
        ws.cell(row=row_cursor, column=8, value="None of these values should be negative")
        row_cursor += 3

        # 10. Table 9: Tax Payable == Paid through cash + Paid through ITC
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_2)")
        ws.cell(row=row_cursor, column=2, value="Total Tax payable")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["tax_payable_T9"])
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_3+4+5+6+7)")
        ws.cell(row=row_cursor, column=2, value="Tax paid through Cash + ITC")
        ws.cell(row=row_cursor, column=3,
                value=valuesFrom9["paid_through_cash_T9"] + valuesFrom9["paid_through_ITC_T9"])
        row_cursor += 1
        ws.cell(row=row_cursor, column=2, value="Tax mismatch")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["tax_payable_T9"] - (
                valuesFrom9["paid_through_cash_T9"] + valuesFrom9["paid_through_ITC_T9"]))
        row_cursor += 3

        # 11. Total tax payable derived from Table 6 of file GSTR-3B_merged.xlsx
        ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_2)")
        ws.cell(row=row_cursor, column=2, value="Total Tax payable (GSTR-9)")
        ws.cell(row=row_cursor, column=3, value=valuesFrom9["tax_payable_T9"])
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T6_SUM(Col_2)")
        ws.cell(row=row_cursor, column=2, value="Total tax payable (GSTR-3B)")
        ws.cell(row=row_cursor, column=3, value=valuesFrom3b["total_tax_payable_column_GSTR_3B_table_6"])
        row_cursor += 1
        ws.cell(row=row_cursor, column=2, value="Tax mismatch")
        ws.cell(row=row_cursor, column=3,
                value=valuesFrom9["tax_payable_T9"] - valuesFrom3b["total_tax_payable_column_GSTR_3B_table_6"])
        row_cursor += 1

        # Apply wrap_text and fixed width to all cells in that column
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 25
            # Apply wrap_text to all cells in that column
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = Alignment(wrap_text=True)

        wb.save(output_dir_of_GSTR_9_report)
        print(f"Excel file saved to: {output_dir_of_GSTR_9_report}")
        print(" === ✅ Returning after successful execution 0f file gstr9_reader.py ===")
    except Exception as e:
        print(f"[GSTR-9 Analysis] ❌ Error during analysis: {e}")


def convert_to_number(value):
    try:
        # Clean value: remove commas, strip spaces
        cleaned = str(value).replace(',', '').strip()
        # Convert to float (handles int too)
        return float(cleaned)
    except (ValueError, TypeError):
        return value  # Leave as-is if not convertible
