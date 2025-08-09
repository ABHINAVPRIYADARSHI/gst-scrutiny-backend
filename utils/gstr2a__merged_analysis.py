import os

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils.globals.constants import result_point_15, result_point_16, empty_string, b2b_merged_sheet, string_N, \
    string_NO, \
    string_YES, string_Y, cdnr_merged_sheet, credit_note, debit_note, isd_merged_sheet, impg_merged_sheet, \
    impg_sez_merged_sheet, eco_merged_sheet, tcs_merged_sheet, tds_merged_sheet

TAX_COL_NAMES = ["Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Cess (₹)"]
ECOM_COL_NAMES = [8, 9, 10, 11, 12]
TCS_COL_NAMES = [5, 6, 7, 8]  # Cess is absent in both TDS & TCS merged tables
TDS_COL_NAMES = [3, 4, 5, 6]
HEADER_ROW = 2
B2B_RATE_COL = 8  # Column I (zero-based)
REVERSE_CHARGE_COL = 7  # Column H
CDNR_REVERSE_CHARGE_COL = 8  # Column I
CDNR_NOTE_TYPE_COL = 2  # Column C
sheet_names = ["B2B_merged", "ISD_merged", "IMPG_merged", "IMPG SEZ_merged"]
EFFECTIVE_DATE_CANCELLATION = 'Effective date of cancellation'


# Old format = up to March 2021,
async def generate_gstr2a_merged_analysis(gstin):
    print(f"[GSTR-2A Analysis] Started execution of method generate_gstr2a_analysis for: {gstin} ===")
    final_result_points = {}
    input_path = f"reports/{gstin}/GSTR-2A_merged.xlsx"
    output_path = f"reports/{gstin}/GSTR-2A_analysis.xlsx"

    try:
        if not os.path.exists(input_path):
            print(f"[GSTR-2A Analysis] Skipped: Input file not found at {input_path}")
            return final_result_points
        reverse_charge_liability_B2B_merged = pd.DataFrame()
        all_sheets = pd.ExcelFile(input_path)
        # Load Excel and extract headers
        if b2b_merged_sheet in all_sheets.sheet_names:
            df_raw = pd.read_excel(input_path, sheet_name=b2b_merged_sheet, header=None)
            # Read data rows after header
            df_B2B_merged = df_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
            # New excel format has multiple header rows. We are setting column names present in 2nd row.
            # So, few columns will have empty names.
            df_B2B_merged.columns = df_raw.iloc[HEADER_ROW]

            # Strip whitespace from all string entries
            df_B2B_merged = df_B2B_merged.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))

            # -- 1. B2B_merged sheet: Reverse Charge Summary --
            df_reverse_B2B = df_B2B_merged[
                df_B2B_merged.iloc[:, REVERSE_CHARGE_COL].astype(str).str.strip().str.upper().isin(
                    [string_Y, string_YES])]
            reverse_charge_liability_B2B_merged = summarize_tax(df_reverse_B2B, TAX_COL_NAMES)
            print("Reverse Charge liability calculation completed.")
        else:
            print("B2B_merged sheet not found.")
        # -- 2. B2B_merged sheet: Cancelled ITC Summary (CANCELLED_DATE_COL should not be empty)--
        # 'Effective date of cancellation' col has d/f positions in old(col S) and new(col U) GSTR-2A excel files.
        # Also, in new format, column name is missing in 3rd header row. So, instead of accessing by position, we will
        # have to access the column by name after setting it in new excel format. Old excel format already has column
        # name, so we can directly access there without any modification.
        try:
            itc_from_cancelled_taxpayers_B2B_merged = pd.DataFrame()
            if df_B2B_merged.shape[1] >= 21:  # the df should have at least 21 columns
                col_name = df_B2B_merged.columns[20]  # store 21st column name in a variable
                # If column name is empty, indicates new excel format, so we set it.
                if not col_name or str(col_name).strip() == empty_string or pd.isna(col_name):
                    columns_list = list(df_B2B_merged.columns)
                    columns_list[20] = EFFECTIVE_DATE_CANCELLATION
                    df_B2B_merged.columns = columns_list
            # Now, access the rows in which 'Effective date of cancellation' col has some date value present
            # itc_from_cancelled_taxpayers_B2B = df_B2B_merged[df_B2B_merged.iloc[:, CANCELLED_DATE_COL].notna()]
            # For old excel formats, EFFECTIVE_DATE_CANCELLATION col name will be auto present
            itc_from_cancelled_taxpayers_B2B = df_B2B_merged[
                df_B2B_merged[EFFECTIVE_DATE_CANCELLATION].notna()  # Not empty
            ]
            print(f"Number of cancelled taxpayers records: {itc_from_cancelled_taxpayers_B2B.shape[0]}")
            itc_from_cancelled_taxpayers_B2B_merged = summarize_tax(itc_from_cancelled_taxpayers_B2B, TAX_COL_NAMES)
            final_result_points["result_point_5_IGST"] = round(itc_from_cancelled_taxpayers_B2B_merged.iloc[0, 0], 2)
            final_result_points["result_point_5_CGST"] = round(itc_from_cancelled_taxpayers_B2B_merged.iloc[0, 1], 2)
            final_result_points["result_point_5_SGST"] = round(itc_from_cancelled_taxpayers_B2B_merged.iloc[0, 2], 2)
            final_result_points["result_point_5_CESS"] = round(itc_from_cancelled_taxpayers_B2B_merged.iloc[0, 3], 2)
            print("Cancelled ITC Summary calculation completed.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing result point 5: {e}")

        # -- 3. ITC by Rate Summary (without grand total row) --
        try:
            # Both new and old formats have column I as 'Rate %', so no issue
            df_rate_grouped = pd.DataFrame()
            df_rate_grouped = (
                df_B2B_merged.groupby(df_B2B_merged.iloc[:, B2B_RATE_COL])[TAX_COL_NAMES]
                    .apply(lambda group: group.apply(pd.to_numeric, errors='coerce').fillna(0).sum())
                    .reset_index()
                    .rename(columns={df_B2B_merged.columns[B2B_RATE_COL]: "Rate %"})
            )
            print("ITC by Rate Summary calculation completed.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing ITC by Rate Summary: {e}")

        # -- 4. Total ITC as per 2A (take data rows from sheets B2B_merged + ISD_merged + IMPG_merged +
        #       IMPG SEZ_merged + Net of CDNR_merged where reverse_charge column = 'N' ). We assume that
        #       3rd row is the header and contains columns "Integrated Tax  (₹)", "Central Tax (₹)",
        #       "State/UT tax (₹)", "Cess (₹)"
        try:
            print("Sheet 'Total ITC as per 2A' calculation started.")
            summary_df = pd.DataFrame()
            summary_rows = []
            for sheet in sheet_names:
                if sheet in all_sheets.sheet_names:
                    print(f"Evaluating sheet: {sheet}")
                    df_raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
                    df = df_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
                    df.columns = df_raw.iloc[HEADER_ROW]
                    # Strip whitespace from all string entries
                    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
                    if sheet == b2b_merged_sheet:
                        df = df[df.iloc[:, REVERSE_CHARGE_COL].astype(str).str.strip().str.upper().isin(
                            [string_N, string_NO])]
                    if not df.empty:
                        tax_sum = []
                        for col in TAX_COL_NAMES:
                            if col in df.columns:
                                col_sum = pd.to_numeric(df[col], errors="coerce").sum(skipna=True)
                            else:
                                print(f"Column: {col} missing in sheet: {sheet}")
                                col_sum = 0
                            tax_sum.append(col_sum)
                    else:
                        tax_sum = [0] * len(TAX_COL_NAMES)
                    # Optionally include the sheet name in the summary row
                    summary_rows.append([sheet] + tax_sum)
                else:
                    print(f"{sheet} sheet not found.")

            # CDNR = Credit Debit Note Regular
            print(f"Evaluating sheet: CDNR_merged")
            cdnr_raw = pd.read_excel(input_path, sheet_name=cdnr_merged_sheet, header=None)
            cdnr_df = cdnr_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
            cdnr_df.columns = cdnr_raw.iloc[HEADER_ROW]
            # Strip whitespace from all string entries
            cdnr_df = cdnr_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
            cdnr_df = cdnr_df[cdnr_df.iloc[:, CDNR_REVERSE_CHARGE_COL].astype(str).str.strip().str.upper().isin(
                [string_N, string_NO])]

            # First append debit note values. Later on append credit note values. Ensure that credit
            # note is the last row to be appended because that is subtracted.
            print(f"Evaluating Debit notes in sheet CDNR_merged")
            debit_df = cdnr_df[cdnr_df.iloc[:, CDNR_NOTE_TYPE_COL].astype(str).str.strip().str.upper() == debit_note]
            if not debit_df.empty:
                tax_sum = debit_df[TAX_COL_NAMES].apply(pd.to_numeric, errors="coerce").sum().tolist()
            else:
                tax_sum = [0] * len(TAX_COL_NAMES)
            summary_rows.append(["Debit Note"] + tax_sum)

            print(f"Evaluating Credit notes in sheet CDNR_merged")
            credit_df = cdnr_df[cdnr_df.iloc[:, CDNR_NOTE_TYPE_COL].astype(str).str.strip().str.upper() == credit_note]
            if not credit_df.empty:
                tax_sum = credit_df[TAX_COL_NAMES].apply(pd.to_numeric, errors="coerce").sum().tolist()
            else:
                tax_sum = [0] * len(TAX_COL_NAMES)
            summary_rows.append(["Credit Note"] + tax_sum)
            summary_df = pd.DataFrame(summary_rows, columns=["Sheet Name", TAX_COL_NAMES[0], TAX_COL_NAMES[1],
                                                             TAX_COL_NAMES[2], TAX_COL_NAMES[3]])

            # debit note values are added, credit note values are subtracted
            # Separate credit note values (last row)
            credit_note_values = summary_df.iloc[-1][TAX_COL_NAMES].apply(pd.to_numeric, errors="coerce")
            # Sum all rows except the last one
            all_except_credit = summary_df.iloc[:-1][TAX_COL_NAMES].apply(pd.to_numeric, errors="coerce").sum()
            # Compute final totals: sum(including debit note) - credit note
            final_total = all_except_credit - credit_note_values
            # Append Total row
            summary_df.loc[len(summary_df.index)] = ["Total"] + final_total.tolist()
            print("Sheet 'Total ITC as per 2A' calculation completed.")

            isd_row = summary_df.loc[summary_df["Sheet Name"] == isd_merged_sheet, TAX_COL_NAMES].iloc[0]
            final_result_points["result_point_20_IGST"] = isd_row.iloc[0]
            final_result_points["result_point_20_CGST"] = isd_row.iloc[1]
            final_result_points["result_point_20_SGST"] = isd_row.iloc[2]
            final_result_points["result_point_20_CESS"] = isd_row.iloc[3]
            print("Result point 20 set.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing result point 20: {e}")

        try:
            impg_total = \
            summary_df.loc[summary_df["Sheet Name"] == impg_merged_sheet, TAX_COL_NAMES].astype(float).iloc[0]
            impg_sez_total = \
            summary_df.loc[summary_df["Sheet Name"] == impg_sez_merged_sheet, TAX_COL_NAMES].astype(float).iloc[0]
            impg_and_impg_sez_total = impg_total + impg_sez_total
            final_result_points["result_point_21_IGST"] = impg_and_impg_sez_total.iloc[0]
            final_result_points["result_point_21_CGST"] = impg_and_impg_sez_total.iloc[1]
            final_result_points["result_point_21_SGST"] = impg_and_impg_sez_total.iloc[2]
            final_result_points["result_point_21_CESS"] = impg_and_impg_sez_total.iloc[3]
            print("Result point 21 set.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing result point 21: {e}")

        # 5. Take full sheet ECOM_merged
        try:
            ecom_df = pd.DataFrame()
            if eco_merged_sheet in all_sheets.sheet_names:
                print("Started processing sheet ECO_merged.")
                ecom_df_raw = pd.read_excel(input_path, sheet_name=eco_merged_sheet, header=None)
                ecom_df = ecom_df_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
                ecom_df.columns = ecom_df_raw.iloc[HEADER_ROW]
                # Strip whitespace from all string entries
                ecom_df = ecom_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
                ecom_df = ecom_df.iloc[:, ECOM_COL_NAMES]
                ecom_df.columns.values[0] = "Taxable Value"
                ecom_df.insert(0, "Label", "")
                ecom_df_total_row = ["TOTAL"]
                for col in ecom_df.columns[1:]:
                    try:
                        total = ecom_df[col].astype(float).sum()
                    except ValueError:
                        total = ""
                    ecom_df_total_row.append(total)
                ecom_df_total_row = pd.DataFrame([ecom_df_total_row], columns=ecom_df.columns)  # Append total row
                ecom_df = pd.concat([ecom_df, ecom_df_total_row], ignore_index=True)
                print("ECOM_merged calculation completed.")
            else:
                print(f"{eco_merged_sheet} sheet not found.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing ECOM_merged: {e}")

        # 6. Take full sheet TCS_merged
        try:
            tcs_df = pd.DataFrame()
            if tcs_merged_sheet in all_sheets.sheet_names:
                tcs_df_raw = pd.read_excel(input_path, sheet_name=tcs_merged_sheet, header=None)
                tcs_df = tcs_df_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
                tcs_df.columns = tcs_df_raw.iloc[HEADER_ROW]
                # Strip whitespace from all string entries
                tcs_df = tcs_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
                tcs_df = tcs_df.iloc[:, TCS_COL_NAMES]
                tcs_df.columns.values[0] = "Taxable Value"
                tcs_df.insert(0, "Label", "")
                tcs_df_total_row = ["TOTAL"]
                for col in tcs_df.columns[1:]:
                    try:
                        total = tcs_df[col].astype(float).sum()
                    except ValueError:
                        total = ""
                    tcs_df_total_row.append(total)
                tcs_df_total_row = pd.DataFrame([tcs_df_total_row], columns=tcs_df.columns)
                tcs_df = pd.concat([tcs_df, tcs_df_total_row], ignore_index=True)
                tcs_taxable_value_total = tcs_df["Taxable Value"].iloc[-1]
                final_result_points[result_point_16] = tcs_taxable_value_total
                print("TCS_merged calculation completed.")
            else:
                print(f"{tcs_merged_sheet} sheet not found.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing TCS_merged: {e}")

        # 7. Take full sheet TDS_merged
        try:
            tds_df = pd.DataFrame()
            if tds_merged_sheet in all_sheets.sheet_names:
                tds_df_raw = pd.read_excel(input_path, sheet_name=tds_merged_sheet, header=None)
                tds_df = tds_df_raw.iloc[HEADER_ROW + 1:].reset_index(drop=True)
                tds_df.columns = tds_df_raw.iloc[HEADER_ROW]
                # Strip whitespace from all string entries
                tds_df = tds_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
                tds_df = tds_df.iloc[:, TDS_COL_NAMES]
                tds_df.columns.values[0] = "Taxable Value"
                tds_df.insert(0, "Label", "")
                tds_df_total_row = ["TOTAL"]
                for col in tds_df.columns[1:]:  # Skip 'Label' column
                    try:
                        total = tds_df[col].astype(float).sum()
                    except ValueError:
                        total = ""  # Leave blank if not numeric
                    tds_df_total_row.append(total)
                tds_df_total_row = pd.DataFrame([tds_df_total_row], columns=tds_df.columns)
                tds_df = pd.concat([tds_df, tds_df_total_row], ignore_index=True)
                tds_taxable_value_total = tds_df["Taxable Value"].iloc[-1]
                final_result_points[result_point_15] = tds_taxable_value_total
                print("TDS_merged calculation completed.")
            else:
                print(f"{tds_merged_sheet} sheet not found.")
        except Exception as e:
            print(f"[GSTR-2A Analysis] Error while computing TDS_merged: {e}")

        # -- Save all summaries to Excel --
        print("Started writing all dataframes to GSTR-2A_analysis.xlsx.")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            if reverse_charge_liability_B2B_merged is not None:
                reverse_charge_liability_B2B_merged.to_excel(writer, index=False, sheet_name="Reverse_Charge_Liability")
            if itc_from_cancelled_taxpayers_B2B_merged is not None:
                itc_from_cancelled_taxpayers_B2B_merged.to_excel(writer, index=False, sheet_name="Cancelled_ITC")
            if df_rate_grouped is not None:
                df_rate_grouped.to_excel(writer, index=False, sheet_name="ITC_by_Rate")
            if summary_df is not None:
                summary_df.to_excel(writer, index=False, sheet_name="Total ITC as per 2A")
            if ecom_df is not None:
                ecom_df.to_excel(writer, index=False, sheet_name="Total ECOM merged")
            if tds_df is not None:
                tds_df.to_excel(writer, index=False, sheet_name="Total TDS merged")
            if tcs_df is not None:
                tcs_df.to_excel(writer, index=False, sheet_name="Total TCS merged")

            # Apply formatting to all written sheets
            for sheet_name, worksheet in writer.sheets.items():
                for col_idx, col in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), start=1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 25  # Set column width
                    for cell in col:
                        cell.alignment = Alignment(wrap_text=True)  # Enable word wrap

        print(f"[GSTR-2A Analysis] ✅ Summary report generated at: {output_path}")
        return final_result_points
    except Exception as e:
        print(f"[GSTR-2A Analysis] ❌ Error during analysis: {e}")
        return final_result_points


def summarize_tax(df: pd.DataFrame, tax_cols):
    """Convert tax columns to numeric and return a one-row summary."""
    df_numeric = df[tax_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
    summary = df_numeric.sum().to_frame().T
    summary["Total Tax"] = summary.sum(axis=1)
    return summary
