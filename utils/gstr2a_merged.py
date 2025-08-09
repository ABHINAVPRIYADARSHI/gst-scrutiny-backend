import os
import pandas as pd
from glob import glob
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import copy

from utils.globals.constants import total_string, sheet_overview

# Define header row ranges per sheet (0-indexed)
header_row_map_new = {
    "B2B": [3, 4, 5],
    "B2BA": [3, 4, 5, 6],
    "CDNR": [3, 4, 5],
    "CDNRA": [3, 4, 5, 6],
    "ECO": [3, 4, 5],
    "ECOA": [3, 4, 5, 6],
    "ISD": [3, 4, 5],
    "ISDA": [3, 4, 5, 6],
    "TDS": [3, 4, 5],
    "TDSA": [3, 4, 5],
    "TCS": [3, 4, 5],
    "IMPG": [3, 4, 5],
    "IMPG SEZ": [3, 4, 5]
}
header_row_map_old = {
    "B2B": [0, 1, 2],
    "B2BA": [0, 1, 2],
    "CDNR": [0, 1, 2],
    "CDNRA": [0, 1, 2],
    # "ECO": [3, 4, 5],
    # "ECOA": [3, 4, 5, 6],
    "ISD": [0, 1, 2],
    "ISDA": [0, 1, 2],
    "TDS": [0, 1, 2],
    "TDSA": [0, 1, 2],
    "TCS": [0, 1, 2],
    "IMPG": [0, 1, 2],
    "IMPGSEZ": [0, 1, 2]
}
# 0-based column indices to apply 'endswith("-Total")' filter
row_filter_column_map = {
    "B2B": 2,  # 3rd column
    "B2BA": 5,  # 6th column
    "CDNR": 3,  # 4th column
    "CDNRA": 3,  # 4th column
    # Add more sheets as needed
}

TAX_HEADER_STANDARDIZATION = {
    "Integrated Tax  (₹)": "Integrated Tax (₹)",
    "Integrated tax (₹)": "Integrated Tax (₹)",
    "State/UT tax (₹)": "State/UT Tax (₹)",
    "State Tax (₹)": "State/UT Tax (₹)",
    "Cess  (₹)": "Cess (₹)",
    "Cess Amount (₹)": "Cess (₹)",
    "Rate(%)": "Rate (%)"
}


def copy_header_with_styles(source_ws, target_ws, header_rows):
    """
    Copies specified header rows from source_ws to target_ws,
    preserving formatting, merged cells, column widths, and row heights.
    """
    header_rows = sorted(header_rows)

    # 1. Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    # 2. Copy styled header rows
    for target_row_idx, source_row_idx in enumerate(header_rows, start=1):
        source_row = source_ws[source_row_idx + 1]  # openpyxl is 1-indexed
        target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[source_row_idx + 1].height

        for col_idx, cell in enumerate(source_row, start=1):
            # new_cell = target_ws.cell(row=target_row_idx, column=col_idx, value=cell.value)
            raw_value = cell.value
            # Normalize header names if they match known messy versions
            if isinstance(raw_value, str) and raw_value.strip() in TAX_HEADER_STANDARDIZATION:
                raw_value = TAX_HEADER_STANDARDIZATION[raw_value.strip()]
            new_cell = target_ws.cell(row=target_row_idx, column=col_idx, value=raw_value)
            if cell.has_style:
                if cell.font: new_cell.font = copy.copy(cell.font)
                if cell.border: new_cell.border = copy.copy(cell.border)
                if cell.fill: new_cell.fill = copy.copy(cell.fill)
                if cell.number_format: new_cell.number_format = cell.number_format
                if cell.protection: new_cell.protection = copy.copy(cell.protection)
                if cell.alignment: new_cell.alignment = copy.copy(cell.alignment)

    # 3. Copy merged cells within header range
    header_row_set = set(r + 1 for r in header_rows)  # 1-based
    for merged_range in source_ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if all(row in header_row_set for row in range(min_row, max_row + 1)):
            try:
                mapped_min_row = header_rows.index(min_row - 1) + 1
                mapped_max_row = header_rows.index(max_row - 1) + 1
                target_ws.merge_cells(
                    start_row=mapped_min_row,
                    start_column=min_col,
                    end_row=mapped_max_row,
                    end_column=max_col
                )
            except ValueError:
                continue


async def generate_gstr2a_merged(input_dir, output_dir):
    print(f"[GSTR-2A_merged] Started execution of method generate_gstr2a_merged for: {input_dir}")

    excel_files = sorted(glob(os.path.join(input_dir, "*.xlsx")))
    if not excel_files:
        raise FileNotFoundError("No Excel files found in the input directory.")

    print(f"[GSTR-2A_merged] Found {len(excel_files)} GSTR-2A Excel files.")
    header_row_map = header_row_map_new  # By-default
    sheet_data = defaultdict(list)
    # readme_copy = None
    # readme_done = False

    # Memory-efficient processing: collect all data first, then concatenate once per sheet
    for file_idx, file_path in enumerate(excel_files):
        try:
            wb = load_workbook(file_path, data_only=True)
            print(f"Processing file: {os.path.basename(file_path)}")

            # Determine format once per file
            current_header_map = header_row_map_old if sheet_overview in wb.sheetnames else header_row_map_new

            for sheet_name in wb.sheetnames:
                if sheet_name not in current_header_map:
                    continue  # Skip unknown sheets early

                ws = wb[sheet_name]
                header_rows = current_header_map[sheet_name]
                data_start_row = max(header_rows) + 1

                # More efficient data extraction - filter empty rows during iteration
                data_rows = [
                    row for row in ws.iter_rows(min_row=data_start_row + 1, values_only=True)
                    if not all(cell is None for cell in row)
                ]

                if not data_rows:
                    sheet_data[sheet_name]  # Ensure sheet exists even if empty
                    continue

                df = pd.DataFrame(data_rows)

                # Apply filtering efficiently
                if (sheet_name in row_filter_column_map and
                        current_header_map == header_row_map_new and not df.empty):
                    col_idx = row_filter_column_map[sheet_name]
                    # More efficient string filtering
                    mask = ~df.iloc[:, col_idx].astype(str).str.endswith(total_string)
                    df = df[mask]

                if not df.empty:
                    sheet_data[sheet_name].append(df)

            wb.close()  # Explicitly close to free memory

        except Exception as e:
            print(f"[GSTR-2A_merged] ❌ Error while copying excel file {file_path}: {str(e)}")
    # Prepare output excel file
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "GSTR-2A_merged.xlsx")
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)

    # Load source workbook once for header copying
    source_wb = None
    try:
        source_wb = load_workbook(excel_files[0], data_only=False)
    except Exception as e:
        print(f"❌ Could not load source file for headers: {str(e)}")
        return None

    # Determine which header map to use for output
    output_header_map = header_row_map_old if sheet_overview in source_wb.sheetnames else header_row_map_new

    for sheet_name in output_header_map:
        try:
            # Create sheet with proper naming
            display_name = "IMPG SEZ" if sheet_name == "IMPGSEZ" else sheet_name
            merged_ws = merged_wb.create_sheet(title=(display_name + "_merged")[:31])

            # Copy formatted header if source sheet exists
            if sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]
                header_rows = output_header_map[sheet_name]
                copy_header_with_styles(source_ws, merged_ws, header_rows)

                # Memory-efficient data writing: single concatenation + bulk write
                df_list = sheet_data.get(sheet_name, [])
                if df_list:
                    print(f"[GSTR-2A_merged] Merging {len(df_list)} DataFrames for sheet {sheet_name}")
                    combined_df = pd.concat(df_list, ignore_index=True)

                    # Clear the list immediately to free memory
                    df_list.clear()

                    # Bulk write data more efficiently
                    data_rows = combined_df.values.tolist()  # Convert to list once
                    for row in data_rows:
                        merged_ws.append(row)

                    # Clear DataFrame to free memory
                    del combined_df
                    print(f"[GSTR-2A_merged] ✅ Written {len(data_rows)} rows to {sheet_name}")
            else:
                print(f"Sheet '{sheet_name}' not found in source. Creating empty sheet.")

        except Exception as e:
            print(f"[GSTR-2A_merged] ❌ Error while writing sheet {sheet_name}: {str(e)}")

    # Close source workbook to free memory
    if source_wb:
        source_wb.close()

    # Clear all sheet data to free memory before saving
    sheet_data.clear()

    merged_wb.save(output_path)
    print(f"✅ [GSTR-2A_merged] merged Excel saved to: {output_path}")
    return output_path
