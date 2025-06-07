import os
import pandas as pd
from glob import glob
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

async def generate_ewb_out_master(input_dir, output_dir):
    excel_files = sorted(glob(os.path.join(input_dir, "*.xlsx")))
    if not excel_files:
        raise FileNotFoundError("No EWB-OUT Excel files found in the input directory.")

    print(f"[EWB-OUT] Found {len(excel_files)} Excel files.")

    sheet_name = None
    header = None
    expected_cols = None
    data_frames = []

    for file_idx, file_path in enumerate(excel_files):
        wb = load_workbook(file_path, data_only=True)
        current_sheet_name = wb.sheetnames[0]  # only one sheet per file

        ws = wb[current_sheet_name]
        print(f"📄 Processing file: {os.path.basename(file_path)}, sheet: {current_sheet_name}")

        # For first file, store sheet name and header row
        if file_idx == 0:
            sheet_name = current_sheet_name
            header = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            expected_cols = len(header)
        # else:
        #     # Validate sheet name matches first file's sheet name
        #     if current_sheet_name != sheet_name:
        #         print(f"⚠️ Sheet name '{current_sheet_name}' does not match first file's sheet name '{sheet_name}'. Skipping file.")
        #         continue

        # Read data rows (starting from row 2)
        data_rows = [
            [cell.value for cell in row]
            for row in ws.iter_rows(min_row=2)
            if not all(cell.value is None for cell in row)
        ]

        if not data_rows:
            print(f"⚠️ No data rows found in file: {os.path.basename(file_path)}")
            continue

        df = pd.DataFrame(data_rows)

        # Validate column count
        actual_cols = df.shape[1]
        if actual_cols != expected_cols:
            print(f"⚠️ Column count mismatch in file '{os.path.basename(file_path)}': expected {expected_cols}, found {actual_cols}. Skipping file.")
            continue

        data_frames.append(df)

    if not data_frames:
        raise ValueError("No valid data found across EWB-OUT files after processing.")

    # Combine all dataframes vertically
    combined_df = pd.concat(data_frames, ignore_index=True)

    # Prepare master workbook and sheet
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "EWB-OUT_Master_Report.xlsx")

    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = (sheet_name + "_merged")[:31]

    # Write header row
    master_ws.append(header)

    # Write combined data rows
    for row in dataframe_to_rows(combined_df, index=False, header=False):
        master_ws.append(row)

    master_wb.save(output_path)
    print(f"✅ EWB-OUT Master Excel saved to: {output_path}")
    return output_path
