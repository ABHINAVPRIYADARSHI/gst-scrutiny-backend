import os
import pandas as pd
from glob import glob
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define header row ranges per sheet (0-indexed)
header_row_map = {
    "B2B": [4, 5, 6],
    "B2BA": [4, 5, 6, 7],
    "CDNR": [4, 5, 6],
    "CDNRA": [4, 5, 6, 7],
    "ECO": [4, 5, 6],
    "ECOA": [4, 5, 6, 7],
    "ISD": [4, 5, 6],
    "ISDA": [4, 5, 6, 7],
    "TDS": [4, 5, 6],
    "TDSA": [4, 5, 6],
    "TCS": [4, 5, 6],
    "IMPG": [4, 5, 6],
    "IMPG SEZ": [4, 5, 6],
}

async def generate_gstr2a_master(input_dir, output_dir):
    excel_files = sorted(glob(os.path.join(input_dir, "*.xlsx")))
    if not excel_files:
        raise FileNotFoundError("No Excel files found in the input directory.")

    print(f"Found {len(excel_files)} GSTR-2A Excel files.")

    sheet_data = defaultdict(list)
    headers_to_copy = {}
    readme_copy = None
    readme_done = False

    for file_idx, file_path in enumerate(excel_files):
        wb = load_workbook(file_path, data_only=True)
        print(f"Processing file: {os.path.basename(file_path)}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if sheet_name.lower().strip() == "read me":
                if not readme_done:
                    readme_copy = ws
                    readme_done = True
                continue

            if sheet_name not in header_row_map:
                print(f"Skipping unknown sheet: {sheet_name}")
                continue

            header_rows = header_row_map[sheet_name]
            data_start_row = max(header_rows) + 1

            # Save plain header values from first file
            if file_idx == 0 and sheet_name not in headers_to_copy:
                headers_to_copy[sheet_name] = [
                    [cell.value for cell in row]
                    for idx, row in enumerate(ws.iter_rows())
                    if idx in header_rows
                ]

            # Extract data rows
            data_rows = []
            for row in ws.iter_rows(min_row=data_start_row + 1, values_only=True):
                if all(cell is None for cell in row):
                    continue
                data_rows.append(row)

            if not data_rows:
                sheet_data[sheet_name]  # ensure key exists
                continue

            df = pd.DataFrame(data_rows)
            sheet_data[sheet_name].append(df)

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "GSTR2A_Master_Report.xlsx")
    master_wb = Workbook()
    master_wb.remove(master_wb.active)

    for sheet_name in header_row_map:
        master_ws = master_wb.create_sheet(title=sheet_name[:31])

        # Paste header rows (just values)
        for row in headers_to_copy.get(sheet_name, []):
            master_ws.append(row)

        # Write stacked data
        df_list = sheet_data.get(sheet_name)
        if df_list:
            combined_df = pd.concat(df_list, ignore_index=True)
            for row in dataframe_to_rows(combined_df, index=False, header=False):
                master_ws.append(row)

    # Write Read me
    if readme_copy:
        readme_ws = master_wb.create_sheet("Read me")
        for row in readme_copy.iter_rows(values_only=True):
            readme_ws.append(row)

    master_wb.save(output_path)
    print(f"✅ GSTR-2A Master Excel saved to: {output_path}")
    return output_path
