from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .gstr3b_merged_reader import gstr3b_merged_reader
from .gstr9_pdf_reader import gstr9_pdf_reader


async def generate_gstr9_Vs_3B_analysis(gstin):
    final_result_points = {}
    valuesFrom9 = {}
    valuesFrom3b ={}
    print(f"[GSTR-9_Vs_3B Analysis] Starting execution of file gstr9_Vs_3B_analysis.py ===")
    output_dir_of_GSTR_9_report = f"reports/{gstin}/GSTR-9 vs GSTR-3B.xlsx"

    try:
        valuesFrom9 = await gstr9_pdf_reader(gstin)
        valuesFrom3b = await gstr3b_merged_reader(gstin)
        # Create a new workbook and sheet (always fresh)
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-9 vs GSTR-3B"
        row_cursor = 1  # Track row position in Excel

        # 1. Table_4_row_G
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T4_G2")
            ws.cell(row=row_cursor, column=2, value=valuesFrom9["table4_G1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table4_G2"]))
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_d")
            ws.cell(row=row_cursor, column=2, value=valuesFrom3b["table_3_1_D1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom3b["table_3_1_D2"]))
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_4_row_G {e}")

        # 2. Table_4_row_N
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T4_N2")
            ws.cell(row=row_cursor, column=2, value=valuesFrom9["table4_N1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table4_N2"]))
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_(cell a + b + d)")
            ws.cell(row=row_cursor, column=3, value=valuesFrom3b["sum_table_3_1_row_a_b_d_taxes"])
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_4_row_N {e}")

        # 3. Table_5_row_D_+_E
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T5_D_+_E")
            ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_5_D1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_5_D2"]))
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_C_+_E")
            ws.cell(row=row_cursor, column=2, value=valuesFrom3b["table_3_1_C1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom3b["table_3_1_C2"]))
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_5_row_D_+_E {e}")

        # 4. Table_5_row_N
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T5_N")
            ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_5_N1"])
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_5_N2"]))
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T3.1_Sum_(a+b+d+e-c)")
            ws.cell(row=row_cursor, column=3, value=valuesFrom3b["sum_table_3_1_A1_B1_D1_E1_minus_C1"])
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_5_row_N {e}")

        # 5. Table_6_row_H
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T6_H (ITC reclaimed)")
            ws.cell(row=row_cursor, column=2, value="Interest Liability on reclaimed ITC: ")
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["table6_row_H_IGST"])
            ws.cell(row=row_cursor, column=4, value=valuesFrom9["table6_row_H_CGST"])
            ws.cell(row=row_cursor, column=5, value=valuesFrom9["table6_row_H_SGST"])
            ws.cell(row=row_cursor, column=6, value=valuesFrom9["table6_row_H_CESS"])
            final_result_points["result_point_11_IGST"] = valuesFrom9["table6_row_H_IGST"]
            final_result_points["result_point_11_CGST"] = valuesFrom9["table6_row_H_CGST"]
            final_result_points["result_point_11_SGST"] = valuesFrom9["table6_row_H_SGST"]
            final_result_points["result_point_11_CESS"] = valuesFrom9["table6_row_H_CESS"]
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_6_row_H {e}")

        # 6. Table_7_row_A
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T7_A (As per Rule 37)")
            ws.cell(row=row_cursor, column=2, value="Interest Liability on reversed ITC: ")
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["table7_row_A_IGST"])
            ws.cell(row=row_cursor, column=4, value=valuesFrom9["table7_row_A_CGST"])
            ws.cell(row=row_cursor, column=5, value=valuesFrom9["table7_row_A_SGST"])
            ws.cell(row=row_cursor, column=6, value=valuesFrom9["table7_row_A_CESS"])
            final_result_points["result_point_10_IGST"] = valuesFrom9["table7_row_A_IGST"]
            final_result_points["result_point_10_CGST"] = valuesFrom9["table7_row_A_CGST"]
            final_result_points["result_point_10_SGST"] = valuesFrom9["table7_row_A_SGST"]
            final_result_points["result_point_10_CESS"] = valuesFrom9["table7_row_A_CESS"]
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_7_row_A {e}")

        # 7. Table_7_row_C
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T7_C (As per Rule 42)")
            ws.cell(row=row_cursor, column=2, value=" Proportionate ITC reversal as per GSTR-9")
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["sum_table7_row_C"])
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged")
            ws.cell(row=row_cursor, column=2, value="Estimate Proportionate reversal as per GSTR-3B")
            ws.cell(row=row_cursor, column=3, value=valuesFrom3b['estimated_ITC_Reversal_IGST'])
            ws.cell(row=row_cursor, column=4, value=valuesFrom3b['estimated_ITC_Reversal_CGST'])
            ws.cell(row=row_cursor, column=5, value=valuesFrom3b['estimated_ITC_Reversal_SGST'])
            ws.cell(row=row_cursor, column=6, value=valuesFrom3b['estimated_ITC_Reversal_CESS'])
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_7_row_C {e}")

        # 8. Table_8_row_D
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T8_D")
            ws.cell(row=row_cursor, column=2, value=convert_to_number(valuesFrom9["table_8_D1"]))
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_8_D4"]))
            ws.cell(row=row_cursor, column=4, value=convert_to_number(valuesFrom9["table_8_D2"]))
            ws.cell(row=row_cursor, column=5, value=convert_to_number(valuesFrom9["table_8_D3"]))
            ws.cell(row=row_cursor, column=6, value=convert_to_number(valuesFrom9["table_8_D5"]))
            ws.cell(row=row_cursor, column=7, value=valuesFrom9["sum_table8_row_D"])
            # if valuesFrom9["sum_table8_row_D"] < floating_zero:
            ws.cell(row=row_cursor, column=8, value="None of these values should be negative. Negative value "
                                                    "signifies excess ITC availed.")
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_8_row_D {e}")

        # 9. Table_8_row_I
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T8_I")
            ws.cell(row=row_cursor, column=2, value=convert_to_number(valuesFrom9["table_8_I1"]))
            ws.cell(row=row_cursor, column=3, value=convert_to_number(valuesFrom9["table_8_I4"]))
            ws.cell(row=row_cursor, column=4, value=convert_to_number(valuesFrom9["table_8_I2"]))
            ws.cell(row=row_cursor, column=5, value=convert_to_number(valuesFrom9["table_8_I3"]))
            ws.cell(row=row_cursor, column=6, value=convert_to_number(valuesFrom9["table_8_I5"]))
            ws.cell(row=row_cursor, column=7, value=valuesFrom9["sum_table8_row_I"])
            ws.cell(row=row_cursor, column=8, value="None of these values should be negative. Negative value "
                                                    "signifies excess credit taken on import than payment made.")
            final_result_points["result_point_18_CGST"] = valuesFrom9["table_8_I2"]
            final_result_points["result_point_18_SGST"] = valuesFrom9["table_8_I3"]
            final_result_points["result_point_18_IGST"] = valuesFrom9["table_8_I4"]
            final_result_points["result_point_18_CESS"] = valuesFrom9["table_8_I5"]
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error Table_8_row_I {e}")

        # 10. Table 9: Tax Payable == Paid through cash + Paid through ITC
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_2)")
            ws.cell(row=row_cursor, column=2, value="Total Tax payable (GSTR-9)")
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["tax_payable_T9"])
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_3+4+5+6+7)")
            ws.cell(row=row_cursor, column=2, value="Tax paid through Cash + ITC")
            ws.cell(row=row_cursor, column=3,
                    value=valuesFrom9["paid_through_cash_T9"] + valuesFrom9["paid_through_ITC_T9"])
            row_cursor += 1
            ws.cell(row=row_cursor, column=2, value="Tax mismatch")
            tax_mismatch = valuesFrom9["tax_payable_T9"] - (
                    valuesFrom9["paid_through_cash_T9"] + valuesFrom9["paid_through_ITC_T9"])
            ws.cell(row=row_cursor, column=3, value=tax_mismatch)
            final_result_points['result_point_17'] = round(tax_mismatch, 2)
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error while Table 9: Tax Payable == Paid through cash + Paid through ITC: {e}")

        # 11. Total tax payable derived from Table 6 of file GSTR-3B_merged.xlsx
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9_T9_SUM(Col_2)")
            ws.cell(row=row_cursor, column=2, value="Total Tax payable (GSTR-9)")
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["tax_payable_table9_IGST"])
            ws.cell(row=row_cursor, column=4, value=valuesFrom9["tax_payable_table9_CGST"])
            ws.cell(row=row_cursor, column=5, value=valuesFrom9["tax_payable_table9_SGST"])
            ws.cell(row=row_cursor, column=6, value=valuesFrom9["tax_payable_table9_CESS"])
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value="GSTR-3B_merged_T6_SUM(Col_2)")
            ws.cell(row=row_cursor, column=2, value="Total tax payable (GSTR-3B)")
            ws.cell(row=row_cursor, column=3, value=valuesFrom3b["table_6_1_total_tax_payable_IGST"])
            ws.cell(row=row_cursor, column=4, value=valuesFrom3b["table_6_1_total_tax_payable_CGST"])
            ws.cell(row=row_cursor, column=5, value=valuesFrom3b["table_6_1_total_tax_payable_SGST"])
            ws.cell(row=row_cursor, column=6, value=valuesFrom3b["table_6_1_total_tax_payable_CESS"])
            row_cursor += 1
            tax_mismatch_point_19_IGST = round(valuesFrom9["tax_payable_table9_IGST"] - valuesFrom3b["table_6_1_total_tax_payable_IGST"], 2)
            tax_mismatch_point_19_CGST = round(valuesFrom9["tax_payable_table9_CGST"] - valuesFrom3b["table_6_1_total_tax_payable_CGST"], 2)
            tax_mismatch_point_19_SGST = round(valuesFrom9["tax_payable_table9_SGST"] - valuesFrom3b["table_6_1_total_tax_payable_SGST"], 2)
            tax_mismatch_point_19_CESS = round(valuesFrom9["tax_payable_table9_CESS"] - valuesFrom3b["table_6_1_total_tax_payable_CESS"], 2)
            ws.cell(row=row_cursor, column=2, value="Tax mismatch")
            ws.cell(row=row_cursor, column=3, value=tax_mismatch_point_19_IGST)
            ws.cell(row=row_cursor, column=4, value=tax_mismatch_point_19_CGST)
            ws.cell(row=row_cursor, column=5, value=tax_mismatch_point_19_SGST)
            ws.cell(row=row_cursor, column=6, value=tax_mismatch_point_19_CESS)
            final_result_points["result_point_19_IGST"] = tax_mismatch_point_19_IGST
            final_result_points["result_point_19_CGST"] = tax_mismatch_point_19_CGST
            final_result_points["result_point_19_SGST"] = tax_mismatch_point_19_SGST
            final_result_points["result_point_19_CESS"] = tax_mismatch_point_19_CESS
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error while Total tax payable derived from Table 6 {e}")

        # Late fee of GSTR-9
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9 late fee")
            if valuesFrom9.get('late_fee_gstr9_applicable') is None:
                print("Setting late_fee_gstr9_applicable from GSTR-3B calculation")
                final_result_points['result_point_13'] = valuesFrom3b.get('result_point_13')
            else:
                ws.cell(row=row_cursor, column=2, value=valuesFrom9["late_fee_gstr9_applicable"])
                final_result_points['result_point_13'] = valuesFrom9['late_fee_gstr9_applicable']
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error while Late fee of GSTR-9 {e}")

        # Result 5.1 = Table 8D + Table 13
        try:
            ws.cell(row=row_cursor, column=1, value="GSTR-9 Table 8D")
            ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_8_D1"])
            ws.cell(row=row_cursor, column=3, value=valuesFrom9["table_8_D4"])
            ws.cell(row=row_cursor, column=4, value=valuesFrom9["table_8_D2"])
            ws.cell(row=row_cursor, column=5, value=valuesFrom9["table_8_D3"])
            ws.cell(row=row_cursor, column=6, value=valuesFrom9["table_8_D5"])
            row_cursor += 1
            # ws.cell(row=row_cursor, column=1, value="GSTR-9 Table 13")
            # ws.cell(row=row_cursor, column=2, value=valuesFrom9["table_13_1"])
            # ws.cell(row=row_cursor, column=3, value=valuesFrom9["table_13_IGST"])
            # ws.cell(row=row_cursor, column=4, value=valuesFrom9["table_13_CGST"])
            # ws.cell(row=row_cursor, column=5, value=valuesFrom9["table_13_SGST"])
            # ws.cell(row=row_cursor, column=6, value=valuesFrom9["table_13_CESS"])
            # row_cursor += 1
            # mismatch_IGST = valuesFrom9["table_8_D4"] + valuesFrom9["table_13_IGST"]
            # mismatch_CGST = valuesFrom9["table_8_D2"] + valuesFrom9["table_13_CGST"]
            # mismatch_SGST = valuesFrom9["table_8_D3"] + valuesFrom9["table_13_SGST"]
            # mismatch_CESS = valuesFrom9["table_8_D5"] + valuesFrom9["table_13_CESS"]
            # ws.cell(row=row_cursor, column=2, value="GSTR-9 (Table 8D + Table 13)")
            # ws.cell(row=row_cursor, column=3, value=mismatch_IGST)
            # ws.cell(row=row_cursor, column=4, value=mismatch_CGST)
            # ws.cell(row=row_cursor, column=5, value=mismatch_SGST)
            # ws.cell(row=row_cursor, column=6, value=mismatch_CESS)

            final_result_points["result_point_5_1_IGST"] = valuesFrom9["table_8_D4"]
            final_result_points["result_point_5_1_CGST"] = valuesFrom9["table_8_D2"]
            final_result_points["result_point_5_1_SGST"] = valuesFrom9["table_8_D3"]
            final_result_points["result_point_5_1_CESS"] = valuesFrom9["table_8_D5"]
            row_cursor += 3
        except Exception as e:
            print(f"[GSTR-9_Vs_3B Analysis] Error while Result 5.1 = Table 8D + Table 13 {e}")

        # Apply wrap_text and fixed width to all cells in that column
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 25
            # Apply wrap_text to all cells in that column
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = Alignment(wrap_text=True)

        wb.save(output_dir_of_GSTR_9_report)
        print(f"Excel file saved to: {output_dir_of_GSTR_9_report}")
        print(" === ✅ Returning after successful execution 0f file gstr9_pdf_reader.py ===")
        return final_result_points
    except Exception as e:
        print(f"[GSTR-9_Vs_3B Analysis] ❌ Error during analysis: {e}")
        return final_result_points


def convert_to_number(value):
    try:
        # Clean value: remove commas, strip spaces
        cleaned = str(value).replace(',', '').strip()
        # Convert to float (handles int too)
        return float(cleaned)
    except (ValueError, TypeError):
        return value  # Leave as-is if not convertible
