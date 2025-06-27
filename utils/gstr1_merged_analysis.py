import pandas as pd
import os

# Constants for fixed column positions
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils.globals.constants import result_point_13, string_yes, string_no, result_point_14

HSN_COL = 0
RATE_COL = 4
TAX_COLS = [5, 6, 7, 8, 9]  # Taxable Value, IGST, CGST, SGST, Cess


async def generate_gstr1_merged_analysis(gstin: str):
    print(f" Started execution of method generate_gstr1_analysis for: {gstin}")
    input_path = f"reports/{gstin}/GSTR-1_merged.xlsx"
    output_path = f"reports/{gstin}/GSTR-1_Analysis.xlsx"
    final_result_points = {}

    if not os.path.exists(input_path):
        print(f"[GSTR-1 Analysis] Skipped: Input file not found at {input_path}")
        return

    try:
        # Read the HSN sheet with header at second row (index 1)
        df = pd.read_excel(input_path, sheet_name="hsn_merged", header=1)

        # Convert relevant tax columns to numeric
        for col in TAX_COLS:
            df.iloc[:, col] = pd.to_numeric(df.iloc[:, col], errors='coerce')

        # Table 1: Group by HSN
        df_by_hsn = df.groupby(df.iloc[:, HSN_COL])[df.columns[TAX_COLS]].sum().reset_index()

        # Table 2: Group by Rate
        df_by_rate = df.groupby(df.iloc[:, RATE_COL])[df.columns[TAX_COLS]].sum().reset_index()

        # Table 3: Group by HSN and Rate using column positions
        df_by_hsn_rate = df.groupby(
            [df.iloc[:, HSN_COL], df.iloc[:, RATE_COL]]
        )[[df.columns[i] for i in TAX_COLS]].sum().reset_index()
        df_by_hsn_rate.columns = ["HSN", "Rate"] + [df.columns[i] for i in TAX_COLS]

        hsn_counts = df_by_hsn_rate.groupby("HSN").size()
        multiple_rows_per_hsn_flag = string_yes if (hsn_counts > 1).any() else string_no
        final_result_points[result_point_14] = multiple_rows_per_hsn_flag

        # Write each table to its own sheet
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_by_hsn.to_excel(writer, index=False, sheet_name="By_HSN")
            df_by_rate.to_excel(writer, index=False, sheet_name="By_Rate")
            df_by_hsn_rate.to_excel(writer, index=False, sheet_name="By_HSN_and_Rate")

            # Add formatting to each sheet
            workbook = writer.book
            for sheet_name in ["By_HSN", "By_Rate", "By_HSN_and_Rate"]:
                worksheet = workbook[sheet_name]

                # Set column widths and enable wrap text
                for col_idx, col in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), start=1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 25
                    for cell in col:
                        cell.alignment = Alignment(wrap_text=True)

        print(f"[GSTR-1 Analysis] ✅ Summary report generated at: {output_path}")
        return final_result_points

    except Exception as e:
        print(f"[GSTR-1 Analysis] ❌ Error during analysis: {e}")
        return final_result_points
