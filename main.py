import os
import glob
import logging
import threading
import webbrowser
import sys
import uvicorn
from typing import List
from fastapi import FastAPI, UploadFile, File, Form, Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from utils.file_handler import save_uploaded_file
from utils.pdf_processor import process_pdf_files
from utils.csv_processor import process_csv_files
from utils.master_generator import generate_merged_excel_and_analysis_report
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from fastapi import HTTPException
from pathlib import Path
import psutil

# Setup logging
logging.basicConfig(level=logging.INFO)

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# Reports directory
REPORTS_BASE_PATH = "reports"


# === Resolve build/static paths (for both normal and exe/frozen) ===
def get_build_path():
    if getattr(sys, 'frozen', False):
        # Running from a PyInstaller .exe
        base_path = sys._MEIPASS
        build_path = os.path.join(base_path, "frontend", "gst-scrutiny-ui", "build")
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
        build_path = os.path.abspath(os.path.join(base_path, "..", "frontend", "gst-scrutiny-ui", "build"))
    return build_path


BUILD_DIR = get_build_path()
STATIC_DIR = os.path.join(BUILD_DIR, "static")

# Serve React static files
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def read_index():
    index_file = os.path.join(BUILD_DIR, "index.html")
    if os.path.exists(index_file):
        return FileResponse(index_file)
    return JSONResponse(content={"error": "Frontend not found"}, status_code=404)


# @app.get("/logo.ico")
# def get_logo():
#     logo_ico_file = os.path.join(BUILD_DIR, "logo.ico")
#     if os.path.exists(logo_ico_file):
#         return FileResponse(logo_ico_file)
#     return JSONResponse(content={"error": "logo.ico not found"}, status_code=404)
#
#
# @app.get("/logo.png")
# def get_logo():
#     logo_png_file = os.path.join(BUILD_DIR, "logo.png")
#     if os.path.exists(logo_png_file):
#         return FileResponse(logo_png_file)
#     return JSONResponse(content={"error": "logo.png not found"}, status_code=404)


# === API Routes ===


@app.post("/upload/")
async def upload_files(
        gstn: str = Form(...),
        return_type: str = Form(...),
        files: List[UploadFile] = File(...)
):
    saved_paths = []
    for file in files:
        path = await save_uploaded_file(file, gstn=gstn, return_type=return_type)
        saved_paths.append(path)
    return {"file_paths": saved_paths}


@app.get("/files/")
def list_uploaded_files(gstn: str, return_type: str):
    folder_path = os.path.join("uploaded_files", gstn, return_type)
    if not os.path.exists(folder_path):
        return {"files": []}

    files = [os.path.basename(path) for path in glob.glob(f"{folder_path}/*")]
    return {"files": files}


@app.post("/process/")
async def process_files(return_type: str, file_paths: List[str]):
    if all(path.endswith(".pdf") for path in file_paths):
        output_path = process_pdf_files(file_paths, return_type)
    elif all(path.endswith(".csv") for path in file_paths):
        output_path = process_csv_files(file_paths, return_type)
    else:
        return {"error": "Mix of CSV and PDF not supported."}
    return {"output": output_path}


# @app.get("/reports/{filename}")
# def get_report_file(filename: str):
#     file_path = os.path.join("reports", filename)
#     return FileResponse(file_path, media_type="application/octet-stream", filename=filename)


@app.delete("/delete/")
def delete_file(gstn: str, return_type: str, filename: str):
    folder_path = os.path.join("uploaded_files", gstn, return_type)
    file_path = os.path.join(folder_path, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        return {"message": "File deleted successfully."}
    return JSONResponse(status_code=404, content={"error": "File not found."})


@app.post("/generate_reports/")
async def generate_master(gstn: str = Form(...),
                          include_ASMT_10_report: str = Form("false")):  # Default to "false" if not provided
    report_flag = include_ASMT_10_report.lower() == "true"
    generated_reports = await generate_merged_excel_and_analysis_report(gstn, report_flag)
    if not generated_reports:
        raise HTTPException(status_code=404, detail="No reports generated for any return type")
    return JSONResponse(content={"status": "completed", "reports": generated_reports})


@app.get("/reports/")
def list_reports(gstn: str = Query(...)):
    reports = os.path.join(REPORTS_BASE_PATH, gstn)
    if not os.path.exists(reports):
        return JSONResponse(status_code=404, content={"detail": "No reports found."})

    files = [f for f in os.listdir(reports)]
    return {"reports": files}


@app.get("/reports/download/")
def download_report(gstn: str = Query(...), filename: str = Query(...)):
    filepath = os.path.join(REPORTS_BASE_PATH, gstn, filename)
    if not os.path.exists(filepath):
        return JSONResponse(status_code=404, content={"detail": "File not found."})
    return FileResponse(path=filepath, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/reports/preview/")
def preview_excel(gstn: str, filename: str):
    file_path = f"reports/{gstn}/{filename}"

    try:
        wb = load_workbook(file_path, data_only=True)
        preview_data = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            preview_data.append({
                "name": sheet,
                "data": rows
            })

        return JSONResponse(content={"sheets": preview_data})

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/check-open-reports/")
def check_open_reports(gstn: str = Query(...)):
    report_path = Path(f"./reports/{gstn}")
    if not report_path.exists():
        return {"open": False}

    for file in report_path.glob("*.xlsx"):
        locked = False
        temp_name = file.with_suffix(".tmp")
        try:
            file.rename(temp_name)
            temp_name.rename(file)  # rename it back
        except PermissionError:
            locked = True
        except Exception:
            continue  # Ignore other errors

        if locked:
            return {"open": True}

    return {"open": False}


# === App Startup ===
if __name__ == "__main__":
    def open_browser():
        webbrowser.open("http://127.0.0.1:8000")


    threading.Timer(1.5, open_browser).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
